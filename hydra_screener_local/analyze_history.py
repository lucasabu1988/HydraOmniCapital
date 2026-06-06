"""
Analizador de Historico del Screener HYDRA Local.

Permite ver el rendimiento de las recomendaciones pasadas.
Incluye:
- Resumen de corridas recientes
- Backtest persistente de las reglas nuevas (composite + short-term boost + strict filter)
  vs el método original (meta_score)
- Integración con Forward Win-Rate Tracking (5d/10d por régimen y Special Mode)
- Export a Excel (--export-excel) que ahora incluye también los últimos runs del historial
  (hojas: Summary, PerDay, RecentRuns, RecentRecommended, Raw)

Uso:
    python analyze_history.py
    python analyze_history.py --days 7 --export-excel
    python analyze_history.py --recompute --no-yf
"""
import json
import os
import argparse
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf
import numpy as np

# Load .env early (centralized) – useful if future features or webhooks are added here
try:
    from utils.env import load_hydra_env
    load_hydra_env()
except Exception as _e:
    print(f"[WARN] .env loader skipped: {_e}")

from core.history import get_recent_runs, list_available_dates
from config import GEOPOLITICAL_RISK_LEVEL, GEO_VOL_THRESHOLD_ADJUST, VOL_SURGE_THRESHOLD, MIN_VOL_THRESHOLD
from core.tracking import aggregate_winrate, print_winrate_report, get_detailed_trades, print_detailed_report

BACKTEST_DIR = "backtest"
BACKTEST_FILE = os.path.join(BACKTEST_DIR, "backtest_results.json")


def show_summary():
    dates = list_available_dates()
    print("\n=== Historico disponible ===")
    print(f"Total de dias guardados: {len(dates)}")
    if dates:
        print(f"Rango: {dates[0]} -> {dates[-1]}")
    print()


def show_last_runs(limit: int = 10):
    runs = get_recent_runs(limit)
    print(f"\n=== Ultimos {len(runs)} dias ===\n")

    for run in runs:
        date = run["date"]
        regime = run.get("regime", {})
        pillars = run.get("pillar_multipliers", {})
        candidates = run.get("top_candidates", [])

        print(f"Date: {date}")
        print(f"   Régimen: {regime.get('score', 0):.2f} ({regime.get('type', '')})")
        if regime.get("special_modes"):
            print(f"   Special Modes: {', '.join(regime['special_modes'])}")

        if pillars:
            print("   Multipliers:", end=" ")
            for p, m in pillars.items():
                print(f"{p}={m:.2f}", end="  ")
            print()

        if candidates:
            recs = [c for c in candidates if c.get("recommended")]
            print(f"   Recomendados ese dia: {len(recs)}")
        print()


def load_backtest_results():
    """Carga el historial acumulado de backtests."""
    if os.path.exists(BACKTEST_FILE):
        with open(BACKTEST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"last_updated": None, "results_by_date": {}, "summary": {}}


def save_backtest_results(results: dict):
    """Guarda los resultados del backtest de forma persistente."""
    os.makedirs(BACKTEST_DIR, exist_ok=True)
    results["last_updated"] = datetime.now().isoformat()
    with open(BACKTEST_FILE, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"  [OK] Backtest guardado en {BACKTEST_FILE}")


def print_backtest_summary(bt):
    """Imprime el resumen acumulado del backtest de forma clara."""
    if not bt or not bt.get("summary"):
        print("No hay datos de backtest disponibles.")
        return
    s = bt["summary"]
    print(f"\nResumen acumulado:")
    print(f"  Días evaluados: {s.get('total_signal_days')}")
    print(f"  Promedio Original (Top10): {s.get('avg_original_1d')}%")
    print(f"  Promedio Nuevas Reglas (Top10): {s.get('avg_new_1d')}%")

    strict = s.get("strict_filter", {})
    if strict.get("avg_1d_when_used") is not None:
        print(f"  Strict Filter - Promedio cuando aplicó: {strict['avg_1d_when_used']}% "
              f"(sobre {strict.get('total_names_passed_across_days', 0)} nombres en {strict.get('days_with_strict_hits', 0)} días)")


def evaluate_signal_day(signal_date: str, forward_days: int = 1):
    """
    Evalúa el rendimiento forward de las recomendaciones de un día.
    Compara:
      - Método original (por meta_score)
      - Método nuevo (por composite_score, si existe en el history)
    """
    history_file = f"history/{signal_date}.json"
    if not os.path.exists(history_file):
        return None

    with open(history_file, "r", encoding="utf-8") as f:
        hist = json.load(f)

    recs = [c for c in hist.get("top_candidates", []) if c.get("recommended")]
    if not recs:
        return None

    tickers = [c["ticker"] for c in recs]

    # Fechas forward
    signal_dt = datetime.strptime(signal_date, "%Y%m%d")
    forward_dates = []
    for i in range(1, forward_days + 1):
        fwd = (signal_dt + timedelta(days=i)).strftime("%Y-%m-%d")
        forward_dates.append(fwd)

    # Descargar precios + volumen (necesario para el strict filter completo)
    try:
        price_data = yf.download(tickers + ["SPY"], start=signal_dt.strftime("%Y-%m-%d"), 
                                  end=(signal_dt + timedelta(days=forward_days+5)).strftime("%Y-%m-%d"),
                                  progress=False, auto_adjust=True)
        if isinstance(price_data.columns, pd.MultiIndex):
            closes = price_data["Close"]
            volumes = price_data["Volume"]
        else:
            closes = price_data
            volumes = None   # fallback, unlikely
    except Exception as e:
        print(f"  Error descargando precios forward para {signal_date}: {e}")
        return None

    results = []
    for c in recs:
        t = c["ticker"]
        try:
            entry_price = float(closes[t].loc[signal_dt.strftime("%Y-%m-%d")])
            fwd_prices = []
            for fd in forward_dates:
                if fd in closes.index.strftime("%Y-%m-%d"):
                    p = float(closes[t].loc[fd])
                    fwd_prices.append((fd, round((p / entry_price - 1)*100, 2)))

            # === Calcular features de corto plazo en el momento de la señal (incluyendo volumen) ===
            series = closes[t].loc[:signal_dt.strftime("%Y-%m-%d")].dropna()
            ret_5d = np.nan
            dist_to_high = np.nan
            vol_ratio = np.nan

            if len(series) >= 25:
                ret_5d = round((series.iloc[-1] / series.iloc[-6] - 1) * 100, 2)
                recent_high = series.iloc[-20:].max()
                dist_to_high = round((series.iloc[-1] / recent_high - 1) * 100, 2)

            # Volumen relativo (últimos 5 días vs promedio 20 días) en la fecha de señal
            if volumes is not None:
                try:
                    vol_series = volumes[t].loc[:signal_dt.strftime("%Y-%m-%d")].dropna()
                    if len(vol_series) >= 25:
                        avg_vol_20 = vol_series.iloc[-20:].mean()
                        avg_vol_5 = vol_series.iloc[-5:].mean()
                        if avg_vol_20 > 0:
                            vol_ratio = round(avg_vol_5 / avg_vol_20, 2)
                except Exception:
                    vol_ratio = np.nan

            if fwd_prices:
                results.append({
                    "ticker": t,
                    "meta_score": c.get("meta_score"),
                    "composite_score": c.get("composite_score", c.get("meta_score")),
                    "short_boost": c.get("short_boost", 0),
                    "ret_5d": ret_5d,
                    "dist_20d_high": dist_to_high,
                    "vol_ratio": vol_ratio,
                    "forward_returns": fwd_prices
                })
        except Exception:
            continue

    if not results:
        return None

    df = pd.DataFrame(results)

    # Selecciones
    top_original = df.sort_values("meta_score", ascending=False).head(10)
    top_new = df.sort_values("composite_score", ascending=False).head(10)

    def avg_return(selection_df, day_offset=0):
        rets = []
        for _, row in selection_df.iterrows():
            fr = row["forward_returns"]
            if len(fr) > day_offset:
                rets.append(fr[day_offset][1])
        return round(np.mean(rets), 2) if rets else None

    # === Strict Filter (las 3 reglas completas con volumen) ===
    # El umbral de volumen se ajusta SOLO según la situación geopolítica
    # NUNCA puede bajar de 1.0 (piso duro)
    dynamic_vol_threshold = VOL_SURGE_THRESHOLD + (GEOPOLITICAL_RISK_LEVEL * GEO_VOL_THRESHOLD_ADJUST)
    dynamic_vol_threshold = max(MIN_VOL_THRESHOLD, dynamic_vol_threshold)
    dynamic_vol_threshold = round(dynamic_vol_threshold, 2)

    strict_mask = (
        (df["ret_5d"] > 15) & 
        (df["dist_20d_high"] >= -2) & 
        (df["vol_ratio"] > dynamic_vol_threshold)
    )
    strict_df = df[strict_mask]

    strict_avg = avg_return(strict_df, 0) if len(strict_df) > 0 else None
    strict_winrate = round((strict_df["forward_returns"].apply(lambda x: x[0][1] > 0).mean() * 100), 0) if len(strict_df) > 0 else None

    return {
        "signal_date": signal_date,
        "n_recommended": len(recs),
        "original_top10_1d_avg": avg_return(top_original, 0),
        "new_top10_1d_avg": avg_return(top_new, 0),
        "geopolitical_risk_level": GEOPOLITICAL_RISK_LEVEL,
        "dynamic_vol_threshold": dynamic_vol_threshold,
        "strict_filter_count": int(len(strict_df)),
        "strict_filter_1d_avg": strict_avg,
        "strict_filter_win_rate": strict_winrate,
        "details": results[:5]
    }


def run_and_save_backtest(recompute: bool = False, max_days: int = None, 
                            forward_days: int = 1, skip_yf: bool = False):
    """
    Ejecuta el backtest sobre el histórico (o parte de él) y guarda resultados.

    Args:
        recompute: Si True, re-evalúa incluso días ya cacheados.
        max_days: Solo procesa los últimos N días de historial.
        forward_days: Cuántos días forward calcular (default 1).
        skip_yf: Si True, no descarga precios forward (solo usa cache o salta evals nuevos).
    """
    print("\n=== Ejecutando Backtest Persistente ===")
    dates = list_available_dates()
    if max_days:
        dates = dates[-max_days:]
    print(f"Fechas de señal encontradas: {dates}")

    current = load_backtest_results()
    results_by_date = current.get("results_by_date", {})

    if recompute:
        print("  --recompute: se ignorará el cache para los días seleccionados.")

    new_evaluations = 0
    for d in dates:
        if not recompute and d in results_by_date:
            continue  # ya evaluado

        if skip_yf:
            if d not in results_by_date:
                print(f"  [skip] {d} ( --no-yf : sin descarga forward )")
            continue

        print(f"  Evaluando {d} (forward +{forward_days}d)...")
        eval_result = evaluate_signal_day(d, forward_days=forward_days)
        if eval_result:
            results_by_date[d] = eval_result
            new_evaluations += 1
            print(f"    +1d Original: {eval_result['original_top10_1d_avg']}% | Nuevo: {eval_result['new_top10_1d_avg']}%")
            if eval_result.get("strict_filter_count", 0) > 0:
                print(f"      Strict Filter: {eval_result['strict_filter_count']} nombres | "
                      f"Vol threshold (geo-adjusted): {eval_result.get('dynamic_vol_threshold')} | "
                      f"+1d: {eval_result.get('strict_filter_1d_avg')}% | "
                      f"WinRate: {eval_result.get('strict_filter_win_rate')}%")

    if new_evaluations > 0:
        # Calcular summary simple
        all_orig = [v["original_top10_1d_avg"] for v in results_by_date.values() if v.get("original_top10_1d_avg") is not None]
        all_new = [v["new_top10_1d_avg"] for v in results_by_date.values() if v.get("new_top10_1d_avg") is not None]

        strict_avgs = [v["strict_filter_1d_avg"] for v in results_by_date.values() 
                       if v.get("strict_filter_1d_avg") is not None]
        strict_counts = [v["strict_filter_count"] for v in results_by_date.values() 
                         if v.get("strict_filter_count") is not None]

        summary = {
            "total_signal_days": len(results_by_date),
            "avg_original_1d": round(np.mean(all_orig), 2) if all_orig else None,
            "avg_new_1d": round(np.mean(all_new), 2) if all_new else None,
            "strict_filter": {
                "avg_1d_when_used": round(np.mean(strict_avgs), 2) if strict_avgs else None,
                "total_names_passed_across_days": int(np.sum(strict_counts)) if strict_counts else 0,
                "days_with_strict_hits": len(strict_avgs)
            },
            "days_evaluated": len(results_by_date)
        }

        final = {
            "last_updated": datetime.now().isoformat(),
            "results_by_date": results_by_date,
            "summary": summary
        }
        save_backtest_results(final)
        print(f"\nBacktest actualizado. Días nuevos evaluados: {new_evaluations}")
    else:
        if skip_yf:
            print("No hay nuevos días para evaluar (--no-yf activo, no se descargaron forwards).")
        else:
            print("No hay nuevos días para evaluar.")

    return load_backtest_results()


def export_backtest_to_excel(bt, output_path: str = None, recent_runs: list = None):
    """
    Exporta los resultados del backtest a un Excel con varias hojas.
    Si se pasa recent_runs (lista de dicts de history), agrega hojas con los últimos runs del historial.
    """
    if not bt:
        print("  No hay datos para exportar.")
        return

    if output_path is None:
        today = datetime.now().strftime("%Y%m%d")
        os.makedirs(BACKTEST_DIR, exist_ok=True)
        output_path = os.path.join(BACKTEST_DIR, f"backtest_analysis_{today}.xlsx")

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Hoja 1: Summary (backtest)
            summary = bt.get("summary", {})
            pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)

            # Hoja 2: Per-day results del backtest (sin details)
            results = bt.get("results_by_date", {})
            if results:
                flat_rows = []
                for date, data in results.items():
                    row = {"date": date}
                    for k, v in data.items():
                        if k != "details" and not isinstance(v, (list, dict)):
                            row[k] = v
                    flat_rows.append(row)
                pd.DataFrame(flat_rows).to_excel(writer, sheet_name="PerDay", index=False)

            # Hoja 3: Raw backtest
            try:
                raw_df = pd.json_normalize(bt)
                raw_df.to_excel(writer, sheet_name="Raw", index=False)
            except Exception:
                pass

            # === Nuevas hojas con los últimos runs del historial ===
            if recent_runs:
                runs_rows = []
                candidates_rows = []

                for run in recent_runs:
                    date = run.get("date", "")
                    regime = run.get("regime", {}) or {}
                    pillars = run.get("pillar_multipliers", {}) or {}
                    candidates = run.get("top_candidates", []) or []
                    recs = [c for c in candidates if c.get("recommended")]

                    runs_rows.append({
                        "date": date,
                        "regime_score": regime.get("score"),
                        "regime_type": regime.get("type"),
                        "special_modes": ", ".join(regime.get("special_modes", [])) if regime.get("special_modes") else "",
                        "num_candidates": len(candidates),
                        "num_recommended": len(recs),
                        **{f"pillar_{k}": round(v, 4) if isinstance(v, (int, float)) else v 
                           for k, v in pillars.items()}
                    })

                    for c in candidates:
                        candidates_rows.append({
                            "date": date,
                            "ticker": c.get("ticker"),
                            "rank": c.get("rank"),
                            "momentum": c.get("momentum"),
                            "meta_score": c.get("meta_score"),
                            "composite_score": c.get("composite_score"),
                            "recommended": bool(c.get("recommended")),
                            "reason": str(c.get("reason", ""))[:60] if c.get("reason") else "",
                            "aggression": c.get("aggression"),
                            "recovery_boost": c.get("recovery_boost"),
                            "ret_5d_10d": c.get("ret_5d_10d"),
                            "dist_20d_high": c.get("dist_20d_high"),
                            "short_boost": c.get("short_boost"),
                            "vol_ratio": c.get("vol_ratio"),
                            "passes_strict": c.get("passes_strict"),
                            "sector": c.get("sector"),
                        })

                if runs_rows:
                    pd.DataFrame(runs_rows).to_excel(writer, sheet_name="RecentRuns", index=False)

                if candidates_rows:
                    pd.DataFrame(candidates_rows).to_excel(writer, sheet_name="RecentRecommended", index=False)

            # ============================================================
            # FORMATO CON OPENPYXL DIRECTO (negritas, anchos, etc.)
            # ============================================================
            try:
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                from openpyxl.utils import get_column_letter

                # Estilos comunes
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style="thin", color="B4B4B4"),
                    right=Side(style="thin", color="B4B4B4"),
                    top=Side(style="thin", color="B4B4B4"),
                    bottom=Side(style="thin", color="B4B4B4")
                )
                data_alignment = Alignment(vertical="center")

                for sheet_name in list(writer.sheets.keys()):
                    ws = writer.sheets[sheet_name]

                    # 1. Formatear encabezado (fila 1)
                    for col_idx, cell in enumerate(ws[1], 1):
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # 2. Congelar fila de encabezado
                    ws.freeze_panes = "A2"

                    # 3. Ajustar anchos de columna automáticamente (con límites razonables)
                    for col_idx, column_cells in enumerate(ws.columns, 1):
                        max_length = 0
                        column_letter = get_column_letter(col_idx)
                        for cell in column_cells:
                            try:
                                if cell.value is not None:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except Exception:
                                pass
                        # Anchos inteligentes: mínimo 10, máximo 45 (excepto para columnas de texto largo)
                        adjusted_width = min(max(max_length + 2, 10), 45)
                        # Hacer un poco más anchas las columnas de texto conocidas
                        if sheet_name in ("RecentRecommended", "PerDay"):
                            if column_letter in ("A", "B", "C"):  # date / ticker / rank suelen ser importantes
                                adjusted_width = min(adjusted_width + 3, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # 4. Bordes y alineación para filas de datos
                    max_row = ws.max_row
                    max_col = ws.max_column
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col), 2):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = data_alignment

                            # Formato numérico para columnas que parecen scores / porcentajes
                            val = cell.value
                            if isinstance(val, (int, float)) and val is not None:
                                if -100 <= val <= 100 and abs(val) < 1000:  # típico de retornos y scores
                                    cell.number_format = '0.00'
                                else:
                                    cell.number_format = '0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")

                    # 5. Ajuste final de altura del header
                    ws.row_dimensions[1].height = 22

            except ImportError:
                # openpyxl está disponible (porque usamos el engine), pero por si acaso
                pass
            except Exception:
                # No romper la exportación por un error de formato
                pass

        print(f"  [OK] Backtest exportado a Excel: {output_path}")
    except Exception as e:
        print(f"  No se pudo exportar a Excel ({e}). Se guardará JSON como fallback.")
        json_path = output_path.replace(".xlsx", ".json") if output_path.endswith(".xlsx") else output_path + ".json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(bt, f, indent=2, ensure_ascii=False, default=str)
        print(f"  [OK] Fallback JSON: {json_path}")


def main():
    parser = argparse.ArgumentParser(
        description="HYDRA Screener - Analizador de Histórico y Backtest Persistente",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python analyze_history.py
  python analyze_history.py --days 5
  python analyze_history.py --recompute --forward-days 1
  python analyze_history.py --no-yf --export-excel
  python analyze_history.py --days 10 --recompute --export-excel
        """
    )
    parser.add_argument("--recompute", action="store_true",
                        help="Re-evaluar todos los días seleccionados (ignora caché del backtest)")
    parser.add_argument("--days", type=int, default=None, metavar="N",
                        help="Solo procesar los últimos N días del historial")
    parser.add_argument("--no-yf", "--no-forward", action="store_true", dest="no_yf",
                        help="No descargar precios con yfinance para cálculos forward (usa solo caché)")
    parser.add_argument("--forward-days", type=int, default=1, metavar="D",
                        help="Número de días forward a evaluar (default: 1)")
    parser.add_argument("--export-excel", action="store_true",
                        help="Exportar a Excel: incluye backtest + últimos runs del historial (hojas RecentRuns y RecentRecommended)")
    parser.add_argument("--last-runs", type=int, default=15, metavar="N",
                        help="Cuántos días recientes mostrar en detalle (default: 15)")

    args = parser.parse_args()

    print("=== HYDRA Screener - Analizador de Historico ===\n")

    # Vistas rápidas del historial
    show_summary()
    recent_limit = args.last_runs or 15
    show_last_runs(limit=recent_limit)
    recent_runs = get_recent_runs(limit=recent_limit)

    # Tracking forward (siempre intenta, es ligero si ya hay datos)
    print("\n=== Forward Win-Rate Tracking (nuevo) ===")
    try:
        report = aggregate_winrate()
        if report and report.get("total_recommendations", 0) > 0:
            print_winrate_report(report)
            trades_df = get_detailed_trades()
            if not trades_df.empty:
                print_detailed_report(trades_df)
        else:
            print("  (Sin datos de tracking forward aún)")
            print("  Ejecuta 'python track_performance.py [--force]' para calcular retornos 5d/10d sobre el historial.")
    except Exception as e:
        print(f"  (No se pudo generar win-rate tracking: {e})")
        print("  Ejecuta 'python track_performance.py' para poblar los datos.")

    # Backtest local (configurable)
    print("\n--- Backtest con reglas nuevas (local) ---")
    bt = run_and_save_backtest(
        recompute=args.recompute,
        max_days=args.days,
        forward_days=args.forward_days,
        skip_yf=args.no_yf
    )

    print_backtest_summary(bt)

    if args.export_excel:
        export_backtest_to_excel(bt, recent_runs=recent_runs)


if __name__ == "__main__":
    main()
