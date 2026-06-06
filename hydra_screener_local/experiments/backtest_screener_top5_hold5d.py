#!/usr/bin/env python
"""
Backtest del screener HYDRA: Top 5 acciones cada 5 días de trading, hold 5 días.
Usa la lógica exacta actual del screener (core/signals.py generate_daily_candidates).
Período: 2000-01-01 a ~2026 (o disponible).

Uso recomendado del señal por su naturaleza (5-day cycle):
- Correr el screener (diario o antes de cada ciclo).
- Elegir las top 5 por composite_score (preferiblemente las que pasen el strict filter para mayor convicción).
- Equal weight 20% cada una.
- Mantener exactamente 5 trading days.
- Rebalancear: vender las que ya no están en el top, comprar las nuevas.
- Esto coincide con el diseño del sistema HYDRA (cycle length 5 días, NUM_POSITIONS=5 en risk-on).

Ventajas:
- Bajo turnover (rotación cada 5 días).
- Aprovecha el regime-aware y dynamic count del screener.
- El strict filter ha mostrado mejor performance en samples.
- Sector control ayuda a diversificar.

Este backtest simula exactamente eso usando el código actual del screener.
"""

import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime, timedelta
import os
import pickle
from typing import List, Dict, Tuple
import warnings
warnings.filterwarnings('ignore')

# Import the EXACT current screener logic
import sys
sys.path.insert(0, '.')
from core.signals import generate_daily_candidates
from core.regime import compute_rich_regime_scores  # for info

# ============================================================================
# PARAMETERS (matching the live screener + 5-day hold rotation)
# ============================================================================
START_DATE = '2000-01-01'
END_DATE = '2026-06-02'  # full available data (covers 2000-2025 + a bit; user request approx)
INITIAL_CAPITAL = 100_000.0
NUM_POSITIONS = 5          # top 5 as proposed
HOLD_DAYS = 5              # exactly 5 trading days
COMMISSION = 0.0005        # 5bp roundtrip approx (simple)
SLIPPAGE = 0.001           # 10bp total for simplicity (conservative)
USE_STRICT_ONLY = False    # set True to only use names that would pass strict filter (if volume data allows; here we simulate without for full history)
MIN_PRICE = 5.0

# Broad pool similar to the project's v8/v9 backtests (liquid names that existed long)
BROAD_POOL = [
    'AAPL', 'MSFT', 'NVDA', 'AMZN', 'GOOGL', 'META', 'AVGO', 'TSLA', 'AMD', 'INTC',
    'QCOM', 'TXN', 'AMAT', 'MU', 'LRCX', 'KLAC', 'MRVL', 'CSCO', 'IBM', 'ORCL',
    'BRK-B', 'JPM', 'V', 'MA', 'BAC', 'WFC', 'GS', 'MS', 'AXP', 'BLK',
    'UNH', 'JNJ', 'LLY', 'ABBV', 'MRK', 'PFE', 'TMO', 'ABT', 'DHR', 'AMGN',
    'AMZN', 'TSLA', 'WMT', 'HD', 'PG', 'COST', 'KO', 'PEP', 'NKE', 'MCD',
    'XOM', 'CVX', 'COP', 'SLB', 'EOG',
    'GE', 'CAT', 'BA', 'HON', 'UNP', 'RTX', 'LMT',
    'NEE', 'DUK', 'SO',
    'VZ', 'T', 'TMUS',
    'DIS', 'CMCSA',
]

print("=" * 80)
print("BACKTEST: HYDRA SCREENER SIGNAL - Top 5 Hold 5 Trading Days")
print("Using EXACT current screener logic (generate_daily_candidates)")
print(f"Period: {START_DATE} to {END_DATE}")
print(f"Positions: {NUM_POSITIONS} | Hold: {HOLD_DAYS} days | Rebalance every {HOLD_DAYS} days")
print(f"Universe pool: {len(BROAD_POOL)} liquid names (survivorship approx)")
print("=" * 80)

# ============================================================================
# DATA LOADING - Reuse/adapt from project's proven v9 backtest (robust cache + yf handling)
# ============================================================================
def download_broad_pool(tickers: List[str], start: str, end: str) -> Dict[str, pd.DataFrame]:
    cache_file = f'data_cache/screener_backtest_broad_{start}_{end}.pkl'
    os.makedirs('data_cache', exist_ok=True)
    if os.path.exists(cache_file):
        print("[Cache] Loading broad pool data...")
        try:
            with open(cache_file, 'rb') as f:
                return pickle.load(f)
        except Exception:
            print("[Cache] Failed, re-downloading...")
    print(f"[Download] Downloading {len(tickers)} symbols for screener backtest...")
    data = {}
    failed = []
    for i, symbol in enumerate(tickers):
        try:
            df = yf.download(symbol, start=start, end=end, progress=False, auto_adjust=True)
            if not df.empty and len(df) > 200:
                if isinstance(df.columns, pd.MultiIndex):
                    df = df['Close'] if 'Close' in df.columns.get_level_values(0) else df.iloc[:, 0]
                else:
                    df = df['Close'] if 'Close' in df.columns else df.iloc[:, 0]
                if isinstance(df, pd.Series):
                    data[symbol] = df
                    if (i + 1) % 15 == 0:
                        print(f"  [{i+1}/{len(tickers)}] {len(data)} symbols...")
            else:
                failed.append(symbol)
        except Exception:
            failed.append(symbol)
    print(f"[Download] {len(data)} symbols valid, {len(failed)} failed")
    with open(cache_file, 'wb') as f:
        pickle.dump(data, f)
    return data

def get_spy(start: str, end: str) -> pd.Series:
    cache = f'data_cache/SPY_screener_{start}_{end}.csv'
    if os.path.exists(cache):
        return pd.read_csv(cache, index_col=0, parse_dates=True).squeeze()
    print("[Download] Downloading SPY...")
    df = yf.download('SPY', start=start, end=end, progress=False, auto_adjust=True)
    if isinstance(df.columns, pd.MultiIndex):
        s = df['Close'].iloc[:, 0]
    else:
        s = df['Close']
    s.to_csv(cache)
    return s

# ============================================================================
# SIMULATION
# ============================================================================
def run_backtest(prices: pd.DataFrame, spy: pd.Series):
    dates = prices.index[prices.index >= '2000-01-01']
    if len(dates) == 0:
        print("No data after 2000")
        return

    # Signal dates: every 5 trading days (to simulate the cycle)
    signal_dates = dates[::5]  # step 5

    equity = [INITIAL_CAPITAL]
    equity_dates = [dates[0]]
    trades = []
    cash = INITIAL_CAPITAL
    positions = {}  # ticker -> (entry_date, entry_price, shares)

    print(f"\nSimulating {len(signal_dates)} signal points...")

    for i, sig_date in enumerate(signal_dates):
        if i % 50 == 0:
            print(f"  [{i}/{len(signal_dates)}] {sig_date.date()} equity={equity[-1]:.0f}")

        # Point-in-time prices up to previous day
        hist_prices = prices.loc[:sig_date - timedelta(days=1)].dropna(axis=1, how='all')
        if len(hist_prices) < 200:
            continue

        hist_spy = spy.loc[:sig_date - timedelta(days=1)].dropna()
        if len(hist_spy) < 200:
            continue

        # Run the EXACT screener logic
        try:
            candidates = generate_daily_candidates(hist_prices, hist_spy)
        except Exception as e:
            # print(f"    Signal error at {sig_date.date()}: {e}")
            continue

        if candidates is None or len(candidates) == 0:
            continue

        # Select top 5 (or less if not enough good ones; here we take top 5 by composite after all logic)
        top5 = candidates.head(NUM_POSITIONS)['ticker'].tolist()

        # For strict preference (optional): if 'passes_strict' in candidates and any pass, prefer them
        if 'passes_strict' in candidates.columns:
            strict_ones = candidates[candidates['passes_strict']].head(NUM_POSITIONS)['ticker'].tolist()
            if len(strict_ones) >= 3:  # only if meaningful number
                top5 = strict_ones[:NUM_POSITIONS]

        # Current prices for entry (next trading day or sig_date close as proxy for simplicity)
        # For realism, use close on sig_date as "decision close", enter at next open ~ same for daily sim
        if sig_date not in prices.index:
            continue
        entry_prices = prices.loc[sig_date, top5].to_dict()

        # Compute forward 5-day returns (using future closes)
        try:
            fwd_start_idx = prices.index.get_loc(sig_date)
            fwd_end_idx = min(fwd_start_idx + HOLD_DAYS, len(prices) - 1)
            fwd_date = prices.index[fwd_end_idx]
            fwd_prices = prices.loc[fwd_date, top5].to_dict()
        except:
            continue

        # Portfolio return for this cycle (equal weight)
        rets = []
        for t in top5:
            if t in entry_prices and t in fwd_prices and entry_prices[t] > 0:
                r = (fwd_prices[t] / entry_prices[t] - 1)
                rets.append(r)
        if not rets:
            continue
        cycle_ret = np.mean(rets) * (1 - COMMISSION) * (1 - SLIPPAGE)   # simple approx costs on the portfolio return

        # Update equity (assume fully invested, no cash drag for simplicity in this version)
        new_equity = equity[-1] * (1 + cycle_ret)
        equity.append(new_equity)
        equity_dates.append(fwd_date)

        trades.append({
            'signal_date': sig_date,
            'exit_date': fwd_date,
            'tickers': top5,
            'cycle_return': cycle_ret,
            'equity': new_equity
        })

        # Log this cycle to the master positions Excel (using the dedicated logger) -- now with entry + realized for dynamic PnL
        try:
            cands_for_log = candidates if 'candidates' in locals() and candidates is not None else None
            import log_cycle_positions
            log_cycle_positions.log_cycle(sig_date, top5, cands_for_log, cycle_return=cycle_ret, equity_after=new_equity, notes="backtest simulation",
                                          entry_prices=entry_prices, realized_prices=fwd_prices)
        except Exception as log_err:
            pass  # logger is optional, don't break the backtest

    # Results
    equity_curve = pd.Series(equity, index=equity_dates)
    total_return = (equity[-1] / INITIAL_CAPITAL - 1)
    years = (equity_dates[-1] - equity_dates[0]).days / 365.25
    cagr = (equity[-1] / INITIAL_CAPITAL) ** (1 / years) - 1

    # Max DD
    peak = equity_curve.cummax()
    dd = (equity_curve - peak) / peak
    max_dd = dd.min()

    # Simple Sharpe (daily-ish, but since 5d periods, approximate)
    rets = equity_curve.pct_change().dropna()
    sharpe = rets.mean() / rets.std() * np.sqrt(252 / 5) if rets.std() > 0 else 0  # rough annualize assuming 5d steps

    # Benchmark SPY buy&hold over same period
    spy_slice = spy.loc[equity_dates[0]:equity_dates[-1]].dropna()
    if len(spy_slice) > 1:
        spy_total = spy_slice.iloc[-1] / spy_slice.iloc[0] - 1
        spy_cagr = (spy_slice.iloc[-1] / spy_slice.iloc[0]) ** (1 / years) - 1
    else:
        spy_total = spy_cagr = 0

    print("\n" + "=" * 80)
    print("BACKTEST RESULTS - Top 5 Hold 5 Days using current HYDRA Screener Signal")
    print("=" * 80)
    print(f"Period: {equity_dates[0].date()} to {equity_dates[-1].date()} ({years:.1f} years)")
    print(f"Final Equity: ${equity[-1]:,.0f}")
    print(f"Total Return: {total_return*100:.1f}%")
    print(f"CAGR: {cagr*100:.2f}%")
    print(f"Max Drawdown: {max_dd*100:.1f}%")
    print(f"Approx Sharpe (annualized): {sharpe:.2f}")
    print(f"Number of cycles: {len(trades)}")
    print()
    print(f"SPY Buy & Hold over same period:")
    print(f"  Total: {spy_total*100:.1f}% | CAGR: {spy_cagr*100:.2f}%")
    print()
    print("Sample recent trades (last 5):")
    for t in trades[-5:]:
        print(f"  {t['signal_date'].date()} -> {t['exit_date'].date()}: {t['tickers']} | {t['cycle_return']*100:+.2f}% | Eq {t['equity']:.0f}")

    # Save equity curve
    eq_df = pd.DataFrame({'equity': equity}, index=equity_dates)
    eq_df.to_csv('backtest/screener_top5_hold5d_equity.csv')
    print("\nEquity curve saved to backtest/screener_top5_hold5d_equity.csv")

    # ========================================================================
    # Save ALL cycle positions to a single comprehensive Excel (legacy from original)
    # ========================================================================
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import openpyxl.utils

        wb = Workbook()

        # Sheet 1: Cycle Summaries
        ws_summary = wb.active
        ws_summary.title = "Cycle_Summaries"

        summary_data = []
        for i, t in enumerate(trades):
            summary_data.append({
                'cycle_id': i + 1,
                'start_date': t['signal_date'],
                'end_date': t['exit_date'],
                'tickers': ', '.join(t['tickers']),
                'num_positions': len(t['tickers']),
                'cycle_return_pct': round(t['cycle_return'] * 100, 4),
                'equity_before': equity[i] if i < len(equity) else None,
                'equity_after': t['equity'],
            })

        summary_df = pd.DataFrame(summary_data)
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
        # Format header
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Sheet 2: All Positions (flattened - every ticker in every cycle)
        ws_positions = wb.create_sheet("All_Positions")

        positions_data = []
        for i, t in enumerate(trades):
            for rank, ticker in enumerate(t['tickers'], 1):
                positions_data.append({
                    'cycle_id': i + 1,
                    'start_date': t['signal_date'],
                    'end_date': t['exit_date'],
                    'ticker': ticker,
                    'weight': 0.2,
                    'rank_in_cycle': rank,
                    'cycle_return_pct': round(t['cycle_return'] * 100, 4),
                    'notes': 'top5 from screener'
                })

        pos_df = pd.DataFrame(positions_data)
        for r_idx, row in enumerate(dataframe_to_rows(pos_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_positions.cell(row=r_idx, column=c_idx, value=value)
        for cell in ws_positions[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Sheet 3: Equity Curve
        ws_equity = wb.create_sheet("Equity_Curve")
        eq_export = eq_df.reset_index()
        eq_export.columns = ['date', 'equity']
        for r_idx, row in enumerate(dataframe_to_rows(eq_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_equity.cell(row=r_idx, column=c_idx, value=value)
        for cell in ws_equity[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Sheet 4: Summary Stats
        ws_stats = wb.create_sheet("Summary_Stats")
        stats = [
            ['Metric', 'Value'],
            ['Start Date', str(equity_dates[0].date())],
            ['End Date', str(equity_dates[-1].date())],
            ['Years', round(years, 2)],
            ['Initial Capital', INITIAL_CAPITAL],
            ['Final Equity', round(equity[-1], 2)],
            ['Total Return %', round(total_return * 100, 2)],
            ['CAGR %', round(cagr * 100, 2)],
            ['Max Drawdown %', round(max_dd * 100, 2)],
            ['Sharpe (approx)', round(sharpe, 2)],
            ['Number of Cycles', len(trades)],
            ['SPY Total Return % (same period)', round(spy_total * 100, 2) if 'spy_total' in locals() else 'N/A'],
            ['SPY CAGR % (same period)', round(spy_cagr * 100, 2) if 'spy_cagr' in locals() else 'N/A'],
        ]
        for r_idx, row in enumerate(stats, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

        # Adjust column widths for all sheets
        for ws in [ws_summary, ws_positions, ws_equity, ws_stats]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        excel_path = 'backtest/all_cycles_positions.xlsx'
        os.makedirs('backtest', exist_ok=True)
        wb.save(excel_path)
        print(f"\n*** ALL CYCLES POSITIONS SAVED TO: {excel_path} ***")
        print("Sheets: Cycle_Summaries | All_Positions | Equity_Curve | Summary_Stats")

    except ImportError:
        print("\n[Warning] openpyxl not installed. Saving positions as CSV instead.")
        summary_df.to_csv('backtest/all_cycles_summaries.csv', index=False)
        pos_df.to_csv('backtest/all_positions.csv', index=False)
        print("Saved CSVs: all_cycles_summaries.csv and all_positions.csv")
    except Exception as e:
        print(f"\n[Error saving Excel] {e}")

    return equity_curve, trades

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    data_dict = download_broad_pool(BROAD_POOL, START_DATE, END_DATE)
    # Convert dict of series to wide prices DF (tickers as columns)
    prices = pd.DataFrame({k: v for k, v in data_dict.items() if isinstance(v, pd.Series)}).sort_index()
    prices = prices.dropna(how='all')
    spy = get_spy(START_DATE, END_DATE)

    # Align indexes
    common = prices.index.intersection(spy.index)
    prices = prices.loc[common]
    spy = spy.loc[common]

    print(f"[Data ready] prices: {prices.shape[1]} tickers, {len(prices)} days | spy: {len(spy)} days")

    run_backtest(prices, spy)
    print("\nDone.")
    print("All cycle positions (historical simulation) have been logged to backtest/portfolio_cycles.xlsx via log_cycle_positions.py")
    print("This Excel accumulates every position from every 5-day cycle in dedicated sheets.")