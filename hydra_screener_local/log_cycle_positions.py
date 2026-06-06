#!/usr/bin/env python
"""
Script to log ALL cycle positions from the HYDRA screener rotation into ONE Excel file.
DYNAMIC PnL TRACKER: each position records entry_price (at signal), current_price (live-updatable),
and PnL is calculated via Excel formulas (=current/entry-1) so the sheet stays dynamic:
- Update current prices (manually in Excel or via refresh_current_prices())
- PnL % and $ auto-recalculate for the 5 actions of each cycle.

Usage:
- For live: after running the screener (with UNIVERSE=all), call log_cycle(signal_date, top5_tickers, candidates_df)
  (entry price auto-fetched via yf as of signal; or pass entry_prices={ticker: price})
- For backtest/historical: pass entry_prices and realized exit as current for accurate historical PnL.

The Excel 'portfolio_cycles.xlsx' (in backtest/) will have:
- Cycle_Summaries: one row per 5-day cycle (+ current cycle PnL snapshot)
- All_Positions: detailed per ticker (rank/score/sector + entry_price, current_price, pnl_pct formula, pnl_usd formula)
- (formulas make it a live portfolio tracker for the 5/5 rotation)

Supports the recommended usage: select top5 (or the dynamic hybrid recommended list sent to Pine/TV), hold exactly 5 trading days, rotate.
The hybrid flow now also logs the *exact* list that was generated for the TradingView dashboard (for precise PnL attribution to what the user actually saw in Pine).
"""

import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import json
import yfinance as yf
import warnings
warnings.filterwarnings('ignore')

EXCEL_PATH = 'backtest/portfolio_cycles.xlsx'
os.makedirs('backtest', exist_ok=True)

def _get_next_trading_day(date):
    """Simple next trading day (Mon-Fri). For production use a proper calendar."""
    d = date
    while True:
        d += timedelta(days=1)
        if d.weekday() < 5:  # 0-4 = Mon-Fri
            return d

def log_cycle(signal_date, tickers, candidates_df=None, cycle_return=None, equity_after=None, notes="",
              entry_prices: dict | None = None, realized_prices: dict | None = None):
    """
    Log one 5-day cycle. Creates/updates the dynamic portfolio_cycles.xlsx with entry/current prices + PnL formulas.

    Args:
        signal_date: date when the screener was run / positions selected
        tickers: list of 5 tickers (the top 5 chosen)
        candidates_df: optional full candidates DataFrame ... (to get rank, score, sector, etc.)
        cycle_return: the realized return for this cycle (if known...)
        equity_after: portfolio equity after this cycle
        notes: any note
        entry_prices: optional dict {ticker: close_price_at_signal} for precise entry (backtest uses sim prices; live auto-fetches if None)
        realized_prices: optional dict {ticker: price} to set as current_price at log time (e.g. backtest fwd/exit price for realized PnL)
    """
    end_date = _get_next_trading_day(_get_next_trading_day(_get_next_trading_day(_get_next_trading_day(_get_next_trading_day(signal_date)))))

    # Load existing or create new
    if os.path.exists(EXCEL_PATH):
        try:
            xls = pd.ExcelFile(EXCEL_PATH)
            summaries = pd.read_excel(xls, sheet_name='Cycle_Summaries') if 'Cycle_Summaries' in xls.sheet_names else pd.DataFrame()
            positions = pd.read_excel(xls, sheet_name='All_Positions') if 'All_Positions' in xls.sheet_names else pd.DataFrame()
        except Exception:
            summaries = pd.DataFrame()
            positions = pd.DataFrame()
    else:
        summaries = pd.DataFrame()
        positions = pd.DataFrame()

    cycle_id = len(summaries) + 1 if not summaries.empty else 1

    # Build summary row (cycle PnL starts at 0; will be snapshot-updated on refresh or when realized)
    summary_row = {
        'cycle_id': cycle_id,
        'start_date': pd.to_datetime(signal_date),
        'end_date': pd.to_datetime(end_date),
        'tickers': ', '.join(tickers),
        'num_positions': len(tickers),
        'cycle_return_pct': round(cycle_return * 100, 4) if cycle_return is not None else None,
        'equity_after': equity_after,
        'notes': notes,
        'cycle_pnl_pct': 0.0,
        'cycle_pnl_usd': 0.0,
    }

    summaries = pd.concat([summaries, pd.DataFrame([summary_row])], ignore_index=True)

    # === PRICE + DYNAMIC PnL LOGIC ===
    # entry: provided or auto-fetch close around signal_date
    # current: realized if provided (backtest), else starts = entry (PnL=0)
    # PnL computed with Excel *formulas* after write for true dynamism (change current_price -> PnL updates)
    NOTIONAL_PER_POSITION = 20_000.0  # assume 100k portfolio, 5x 20% equal weight for $ PnL tracking

    def _fetch_price_asof(tkr: str, asof: datetime, lookback_days: int = 7) -> float | None:
        """Fetch adjusted close on or before asof (for entry at signal or historical fill)."""
        try:
            start = (asof - timedelta(days=lookback_days)).date()
            end = (asof + timedelta(days=2)).date()
            dfp = yf.download(tkr, start=str(start), end=str(end), progress=False, auto_adjust=True)
            if dfp is not None and not dfp.empty:
                closes = dfp['Close'] if isinstance(dfp.columns, pd.MultiIndex) else dfp['Close']
                if isinstance(closes, pd.DataFrame):
                    closes = closes.iloc[:, 0]
                # last available <= asof
                idx = closes.index[closes.index <= pd.to_datetime(asof)]
                if len(idx) > 0:
                    return float(closes.loc[idx[-1]])
                return float(closes.iloc[-1])
        except Exception:
            pass
        return None

    # Build positions rows (one per ticker)
    pos_rows = []
    for rank, ticker in enumerate(tickers, 1):
        row = {
            'cycle_id': cycle_id,
            'start_date': pd.to_datetime(signal_date),
            'end_date': pd.to_datetime(end_date),
            'ticker': ticker,
            'weight': 0.20,
            'rank_in_cycle': rank,
        }

        if candidates_df is not None and not candidates_df.empty:
            match = candidates_df[candidates_df['ticker'] == ticker]
            if not match.empty:
                m = match.iloc[0]
                row['composite_score'] = m.get('composite_score')
                row['ret_5d_10d'] = m.get('ret_5d_10d')
                row['sector'] = m.get('sector')
                row['passes_strict'] = m.get('passes_strict')
                row['sector_penalty_applied'] = m.get('sector_penalty_applied')
                row['rank_at_signal'] = m.get('rank')  # overall rank in the big combined list

        # Entry price (precise if passed from backtest/sim, else fetch asof signal)
        ep = entry_prices
        if ep is not None and hasattr(ep, 'to_dict'):
            ep = ep.to_dict()
        entry_p = None
        if ep and ticker in ep and ep[ticker]:
            entry_p = float(ep[ticker])
        else:
            entry_p = _fetch_price_asof(ticker, pd.to_datetime(signal_date))

        row['entry_price'] = round(entry_p, 4) if entry_p and entry_p > 0 else None

        # Current price: if realized provided (backtest exit), use it; else start at entry (PnL 0 until refresh or manual)
        rp = realized_prices
        if rp is not None and hasattr(rp, 'to_dict'):
            rp = rp.to_dict()
        curr_p = entry_p
        if rp and ticker in rp and rp[ticker]:
            curr_p = float(rp[ticker])
        row['current_price'] = round(curr_p, 4) if curr_p and curr_p > 0 else (row['entry_price'])

        # Initial static (will be replaced by formula cells below for dynamism)
        row['pnl_pct'] = 0.0
        row['pnl_usd'] = 0.0

        pos_rows.append(row)

    positions = pd.concat([positions, pd.DataFrame(pos_rows)], ignore_index=True)

    # Post-concat migration + canonical column order (prices/PnL at end for nice Excel; old rows get NaN for new cols)
    base_pos_order = ['cycle_id', 'start_date', 'end_date', 'ticker', 'weight', 'rank_in_cycle',
                      'composite_score', 'ret_5d_10d', 'sector', 'passes_strict',
                      'sector_penalty_applied', 'rank_at_signal']
    dyn_pos_order = ['entry_price', 'current_price', 'pnl_pct', 'pnl_usd']
    for c in base_pos_order + dyn_pos_order:
        if c not in positions.columns:
            positions[c] = None
    other_pos = [c for c in positions.columns if c not in (base_pos_order + dyn_pos_order)]
    positions = positions[base_pos_order + dyn_pos_order + other_pos]

    # Same for summaries (cycle PnL cols at end)
    base_sum_order = ['cycle_id', 'start_date', 'end_date', 'tickers', 'num_positions',
                      'cycle_return_pct', 'equity_after', 'notes']
    dyn_sum_order = ['cycle_pnl_pct', 'cycle_pnl_usd']
    for c in base_sum_order + dyn_sum_order:
        if c not in summaries.columns:
            summaries[c] = None
    other_sum = [c for c in summaries.columns if c not in (base_sum_order + dyn_sum_order)]
    summaries = summaries[base_sum_order + dyn_sum_order + other_sum]

    # Write back to Excel with formatting + **Excel formulas for PnL** (the key for "excel dinámico")
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        summaries.to_excel(writer, sheet_name='Cycle_Summaries', index=False)
        positions.to_excel(writer, sheet_name='All_Positions', index=False)

        wb = writer.book

        # === All_Positions: locate price cols and inject FORMULAS for pnl_pct / pnl_usd ===
        if 'All_Positions' in wb.sheetnames:
            ws = wb['All_Positions']
            headers = [cell.value for cell in ws[1]]
            # Expected new cols at end (after migration/append)
            try:
                entry_idx = headers.index('entry_price') + 1
                curr_idx = headers.index('current_price') + 1
                pnl_pct_idx = headers.index('pnl_pct') + 1
                pnl_usd_idx = headers.index('pnl_usd') + 1
                entry_col = get_column_letter(entry_idx)
                curr_col = get_column_letter(curr_idx)
                pnl_pct_col = get_column_letter(pnl_pct_idx)
                pnl_usd_col = get_column_letter(pnl_usd_idx)
            except ValueError:
                entry_col = curr_col = pnl_pct_col = pnl_usd_col = None

            # Apply header style (extended)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

            # Set formulas + number formats for data rows
            for r in range(2, ws.max_row + 1):
                if entry_col and curr_col and pnl_pct_col:
                    # Formula: if valid prices, (current/entry - 1), else blank/0
                    ws[f'{pnl_pct_col}{r}'] = f'=IF(AND(ISNUMBER({curr_col}{r}),ISNUMBER({entry_col}{r}),{entry_col}{r}<>0),{curr_col}{r}/{entry_col}{r}-1,0)'
                    ws[f'{pnl_pct_col}{r}'].number_format = '0.00%'
                if pnl_usd_col and pnl_pct_col:
                    ws[f'{pnl_usd_col}{r}'] = f'={pnl_pct_col}{r}*{NOTIONAL_PER_POSITION:.0f}'
                    ws[f'{pnl_usd_col}{r}'].number_format = '$#,##0.00'
                # price formats
                if entry_col:
                    ws[f'{entry_col}{r}'].number_format = '#,##0.00'
                if curr_col:
                    ws[f'{curr_col}{r}'].number_format = '#,##0.00'

            # Auto width (improved for new cols)
            for col_cells in ws.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 42)

        # === Cycle_Summaries header + basic ===
        if 'Cycle_Summaries' in wb.sheetnames:
            ws_sum = wb['Cycle_Summaries']
            for cell in ws_sum[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col_cells in ws_sum.columns:
                max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
                ws_sum.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 42)

    print(f"[Cycle Logger] Logged cycle {cycle_id} | {pd.to_datetime(signal_date).date()} -> {pd.to_datetime(end_date).date()} | {tickers}")
    return cycle_id


def _fetch_price_asof_for_refresh(tkr: str, asof: datetime, lookback_days: int = 7) -> float | None:
    """Small helper for backfill during refresh (and available for log inner)."""
    try:
        start = (asof - timedelta(days=lookback_days)).date()
        end = (asof + timedelta(days=2)).date()
        dfp = yf.download(tkr, start=str(start), end=str(end), progress=False, auto_adjust=True)
        if dfp is not None and not dfp.empty:
            closes = dfp['Close'] if isinstance(dfp.columns, pd.MultiIndex) else dfp['Close']
            if isinstance(closes, pd.DataFrame):
                closes = closes.iloc[:, 0]
            idx = closes.index[closes.index <= pd.to_datetime(asof)]
            if len(idx) > 0:
                return float(closes.loc[idx[-1]])
            return float(closes.iloc[-1])
    except Exception:
        pass
    return None


# ============================================================================
# DYNAMIC UPDATE: refresh current prices + cycle PnL snapshots (the "dinámico" part)
# ============================================================================
NOTIONAL_PER_POSITION = 20_000.0  # keep in sync with log for $ PnL

def refresh_current_prices(lookback_cycles: int | None = None, excel_path: str = None) -> int:
    """
    Make the Excel dynamic: fetch latest market closes for positions and update 'current_price'.
    - PnL formulas in All_Positions will reflect new values when opened in Excel.
    - Also updates Cycle_Summaries 'cycle_pnl_pct' / 'cycle_pnl_usd' as python-computed snapshot (avg/sum of the cycle's 5 positions).
    - If lookback_cycles is int, only refresh positions from the last N cycles (faster for large history).
    - Safe for old rows without prices (tries to backfill entry too).
    Returns: number of position rows updated with a new current price.
    """
    path = excel_path or EXCEL_PATH
    if not os.path.exists(path):
        print(f"[refresh] No Excel at {path}")
        return 0

    wb = load_workbook(path)
    if 'All_Positions' not in wb.sheetnames:
        print("[refresh] No All_Positions sheet")
        return 0

    ws = wb['All_Positions']
    headers = [c.value for c in ws[1]]
    # col indices (1-based)
    try:
        cycle_col = headers.index('cycle_id') + 1
        ticker_col = headers.index('ticker') + 1
        entry_col = headers.index('entry_price') + 1
        curr_col = headers.index('current_price') + 1
        start_col = headers.index('start_date') + 1
    except ValueError as e:
        print(f"[refresh] Missing required columns: {e}")
        return 0

    # Collect rows to potentially update
    rows_to_update = []  # (row_num, cycle_id, ticker, entry_val_or_None)
    max_cycle = 0
    for r in range(2, ws.max_row + 1):
        cyc = ws.cell(row=r, column=cycle_col).value
        tkr = ws.cell(row=r, column=ticker_col).value
        if cyc is not None:
            max_cycle = max(max_cycle, int(cyc))
        if tkr:
            rows_to_update.append((r, cyc, str(tkr), ws.cell(row=r, column=entry_col).value))

    if lookback_cycles is not None and max_cycle > 0:
        min_cyc = max_cycle - lookback_cycles + 1
        rows_to_update = [x for x in rows_to_update if x[1] is None or int(x[1]) >= min_cyc]

    unique_tickers = sorted(set(x[2] for x in rows_to_update if x[2]))
    if not unique_tickers:
        print("[refresh] No tickers to update")
        return 0

    print(f"[refresh] Fetching latest prices for {len(unique_tickers)} tickers (lookback_cycles={lookback_cycles}) ...")
    latest = {}
    try:
        # Batch fetch recent data (auto_adjust for splits)
        df_live = yf.download(unique_tickers, period='10d', progress=False, auto_adjust=True, group_by='ticker')
        if df_live is not None and not df_live.empty:
            for tkr in unique_tickers:
                try:
                    if isinstance(df_live.columns, pd.MultiIndex):
                        sub = df_live[tkr]['Close'].dropna()
                    else:
                        sub = df_live['Close'].dropna() if len(unique_tickers) == 1 else df_live[tkr].dropna() if tkr in df_live else pd.Series()
                    if len(sub) > 0:
                        latest[tkr] = float(sub.iloc[-1])
                except Exception:
                    pass
    except Exception as e:
        print(f"[refresh] yf batch failed: {e}")

    # Fallback per-ticker for any missing
    for tkr in unique_tickers:
        if tkr not in latest:
            p = None
            try:
                d = yf.download(tkr, period='5d', progress=False, auto_adjust=True)
                if d is not None and not d.empty:
                    cl = d['Close'] if not isinstance(d.columns, pd.MultiIndex) else d['Close'].iloc[:, 0]
                    p = float(cl.dropna().iloc[-1])
            except:
                pass
            if p:
                latest[tkr] = p

    updated = 0
    cycle_latest_pnls: dict[int, list] = {}  # for summary agg

    for (r, cyc, tkr, entry_val) in rows_to_update:
        new_curr = latest.get(tkr)
        if new_curr is None or new_curr <= 0:
            continue

        # Backfill entry if missing (for pre-dynamic cycles or bad fetch at log)
        if entry_val is None or entry_val <= 0:
            # try historical asof start_date of this row
            try:
                sdate = ws.cell(row=r, column=start_col).value
                if sdate:
                    entry_val = _fetch_price_asof_for_refresh(tkr, pd.to_datetime(sdate))  # reuse helper? inline simple
                    if entry_val:
                        ws.cell(row=r, column=entry_col).value = round(entry_val, 4)
                        ws.cell(row=r, column=entry_col).number_format = '#,##0.00'
            except:
                pass

        entry_val = entry_val or ws.cell(row=r, column=entry_col).value
        ws.cell(row=r, column=curr_col).value = round(new_curr, 4)
        ws.cell(row=r, column=curr_col).number_format = '#,##0.00'

        # Compute py pnl for this row (to feed cycle aggregate; formula in file will match when opened)
        if entry_val and entry_val > 0:
            pnl = (new_curr / entry_val - 1)
            cyc_int = int(cyc) if cyc is not None else 0
            cycle_latest_pnls.setdefault(cyc_int, []).append(pnl)
        updated += 1

    # Update Cycle_Summaries snapshots (last refreshed PnL)
    if 'Cycle_Summaries' in wb.sheetnames and cycle_latest_pnls:
        ws_sum = wb['Cycle_Summaries']
        sum_headers = [c.value for c in ws_sum[1]]
        try:
            sum_cyc_idx = sum_headers.index('cycle_id') + 1
            sum_pnl_pct_idx = sum_headers.index('cycle_pnl_pct') + 1 if 'cycle_pnl_pct' in sum_headers else None
            sum_pnl_usd_idx = sum_headers.index('cycle_pnl_usd') + 1 if 'cycle_pnl_usd' in sum_headers else None
        except ValueError:
            sum_pnl_pct_idx = sum_pnl_usd_idx = None

        if sum_pnl_pct_idx:
            for r in range(2, ws_sum.max_row + 1):
                cyc = ws_sum.cell(row=r, column=sum_cyc_idx).value
                if cyc is None:
                    continue
                pnls = cycle_latest_pnls.get(int(cyc), [])
                if pnls:
                    avg = sum(pnls) / len(pnls)
                    ws_sum.cell(row=r, column=sum_pnl_pct_idx).value = round(avg, 6)
                    ws_sum.cell(row=r, column=sum_pnl_pct_idx).number_format = '0.00%'
                    if sum_pnl_usd_idx:
                        ws_sum.cell(row=r, column=sum_pnl_usd_idx).value = round(avg * NOTIONAL_PER_POSITION * len(pnls), 2)
                        ws_sum.cell(row=r, column=sum_pnl_usd_idx).number_format = '$#,##0.00'

    wb.save(path)
    print(f"[refresh] Updated current_price for {updated} position rows. Cycle PnL snapshots refreshed.")
    print(f"         File: {path}")
    print("         Open in Excel: PnL formulas will calculate live from the updated current_price values.")
    return updated


# ============================================================================
# DEMO / BACKTEST INTEGRATION EXAMPLE
# ============================================================================
if __name__ == "__main__":
    print("=== Cycle Positions Logger Demo (DYNAMIC PnL) ===")

    # Example 1: Log the most recent real cycle from the screener run with "all"
    try:
        with open('history/20260601.json') as f:
            hist = json.load(f)
        recs = [c for c in hist.get('top_candidates', []) if c.get('recommended')][:5]
        top5 = [c['ticker'] for c in recs]
        cands = pd.DataFrame(recs)
        log_cycle(pd.to_datetime('2026-06-01'), top5, cands, notes="Latest real run with UNIVERSE=all")
    except Exception as e:
        print(f"Could not load latest history (ok if no file): {e}")

    # Example 2: Simulate a few historical cycles (for demo) -- with fake entry for illustration
    import datetime as dt
    dummy_dates = [dt.date(2026,5,20), dt.date(2026,5,27), dt.date(2026,6,3)]
    dummy_tops = [
        ['DELL', 'ARM', 'MRVL', 'HPE', 'STX'],
        ['DELL', 'MU', 'HPE', 'STX', 'NTAP'],
        ['ARM', 'DELL', 'MRVL', 'CRWD', 'FTNT'],
    ]
    for d, tops in zip(dummy_dates, dummy_tops):
        # demo: pass dummy entry prices (in real backtest use real hist prices)
        fake_entries = {t: 100.0 + i*2 for i, t in enumerate(tops)}
        log_cycle(pd.to_datetime(d), tops, notes="Simulated historical cycle", entry_prices=fake_entries)

    print(f"\nExcel saved to: {EXCEL_PATH}")

    # Demo the dynamic refresh (will set current to latest market for the tickers, update PnL snapshots)
    try:
        n = refresh_current_prices(lookback_cycles=5)
        print(f"Refresh demo updated {n} rows.")
    except Exception as ex:
        print(f"Refresh demo skipped (no net/yf?): {ex}")

    print("\nOpen the Excel: it now has entry_price + current_price columns.")
    print("PnL % and PnL $ are FORMULAS -- edit any current_price or let refresh update, and PnL recalcs automatically in Excel for the 5 per cycle.")
    print("Cycle_Summaries has cycle_pnl_pct/usd snapshots from last refresh.")